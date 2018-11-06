import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os
import json
import keras
import pandas as pd
import numpy as np
from keras.models import model_from_json
from sklearn import preprocessing
from sklearn.externals import joblib
from datetime import datetime
import pickle
from openpyxl import load_workbook
pd.options.display.max_columns = 500
pd.options.display.max_rows = 1000
from keras import backend as K

# Функции для заполнения пропусков
# Если пропущен химиечский элемент - добавляеться среднее значение
# Если пропущено значение ОКБ отпуска или ОКБ закалки - устанавливаеться 3 (самое популярное ОКБ)
# Остальные пропуски заполняються нулями, чтобы в дальнейшем отправить данные строки в ошибочный файл

def auto_fill_C(x):
    if (x == 0) | (x == None) | (pd.isnull(x)):
        return 0.15
    return x

def auto_fill_Mn(x):
    if (x == 0) | (x == None)| (pd.isnull(x)):
        return 0.55
    return x

def auto_fill_Si(x):
    if (x == 0) | (x == None)| (pd.isnull(x)):
        return 0.26
    return x

def auto_fill_Cr(x):
    if (x == 0) | (x == None)| (pd.isnull(x)):
        return 0.55
    return x

def auto_fill_Ni(x):
    if (x == 0) | (x == None)| (pd.isnull(x)):
        return 0.13
    return x

def auto_fill_Cu(x):
    if (x == 0) | (x == None)| (pd.isnull(x)):
        return 0.22
    return x

def auto_fill_Al(x):
    if (x == 0) | (x == None)| (pd.isnull(x)):
        return 0.03
    return x

def auto_fill_okb(x):
    if pd.isnull(x):
        return 3
    return x

def auto_fill_null(x):
    if pd.isnull(x):
        return 0
    return x

def cl_inf(x):
    if x==float('inf'):
        x = 0
    return x

# Функция вызывающая функции для заполнения пропусков в input файле

def auto_fill_empty(df):
    df[u'C'] = df[u'C'].apply(auto_fill_C)
    df[u'Mn'] = df[u'Mn'].apply(auto_fill_Mn)
    df[u'Si'] = df[u'Si'].apply(auto_fill_Si)
    df[u'Cr'] = df[u'Cr'].apply(auto_fill_Cr)
    df[u'Ni'] = df[u'Ni'].apply(auto_fill_Ni)
    df[u'Cu'] = df[u'Cu'].apply(auto_fill_Cu)
    df[u'Al'] = df[u'Al'].apply(auto_fill_Al)
    df[u'ОКБ (отпуск)'] = df[u'ОКБ (отпуск)'].apply(auto_fill_okb)
    df[u'ОКБ (закалка)'] = df[u'ОКБ (закалка)'].apply(auto_fill_okb)
    df[u'диаметр'] = df[u'диаметр'].apply(auto_fill_null)
    df[u'толщина стенки'] = df[u'толщина стенки'].apply(auto_fill_null)
    df[u'темп-ра терм. (норм.)'] = df[u'темп-ра терм. (норм.)'].apply(auto_fill_null)
    df[u'темп-ра спрейр (норм.)'] = df[u'темп-ра спрейр (норм.)'].apply(auto_fill_null)
    df[u'скорость движения (норм.)'] = df[u'скорость движения (норм.)'].apply(auto_fill_null)
    df[u'темп-ра терм. (отпуск)'] = df[u'темп-ра терм. (отпуск)'].apply(auto_fill_null)
    df[u'темп-ра спрейр (отпуск)'] = df[u'темп-ра спрейр (отпуск)'].apply(auto_fill_null)
    df[u'скорость движения (отпуск)'] = df[u'скорость движения (отпуск)'].apply(auto_fill_null)
    df[u'расход воды (норм.) (1)'] = df[u'расход воды (норм.) (1)'].apply(auto_fill_null)
    df[u'расход воды (норм.) (2)'] = df[u'расход воды (норм.) (2)'].apply(auto_fill_null)
    df[u'расход воды (норм.) (3)'] = df[u'расход воды (норм.) (3)'].apply(auto_fill_null)
    return df


# Функция, расчитывающая признаки, которых нет в input файле, но используються в моделе

def add_len_and_param(data):
    data.loc[data[u'ОКБ (отпуск)'] > 4, u'ОКБ (отпуск)'] = 3
    data.loc[data[u'ОКБ (закалка)'] > 4, u'ОКБ (закалка)'] = 3
    data[u'длина (отпуск)'] = None
    data.loc[data[u'ОКБ (отпуск)'] == 1.0, u'длина (отпуск)'] = 2.2
    data.loc[data[u'ОКБ (отпуск)'] == 2.0, u'длина (отпуск)'] = 2.2
    data.loc[data[u'ОКБ (отпуск)'] == 3.0, u'длина (отпуск)'] = 1.2
    data.loc[data[u'ОКБ (отпуск)'] == 4.0, u'длина (отпуск)'] = 2.9
    data[u'длина (закалка)'] = None
    data.loc[data[u'ОКБ (закалка)'] == 1.0, u'длина (закалка)'] = 2.2
    data.loc[data[u'ОКБ (закалка)'] == 2.0, u'длина (закалка)'] = 2.2
    data.loc[data[u'ОКБ (закалка)'] == 3.0, u'длина (закалка)'] = 2.2
    data.loc[data[u'ОКБ (закалка)'] == 4.0, u'длина (закалка)'] = 1.65

    data[u'норм_1'] = 0
    data[u'норм_2'] = 0
    data[u'норм_3'] = 0
    data[u'норм_4'] = 0
    data[u'отпуск_1'] = 0
    data[u'отпуск_2'] = 0
    data[u'отпуск_3'] = 0
    data[u'отпуск_4'] = 0

    data.loc[data[u'ОКБ (отпуск)'] == 1.0, u'отпуск_1'] = 1
    data.loc[data[u'ОКБ (отпуск)'] == 2.0, u'отпуск_2'] = 1
    data.loc[data[u'ОКБ (отпуск)'] == 3.0, u'отпуск_3'] = 1
    data.loc[data[u'ОКБ (отпуск)'] == 4.0, u'отпуск_4'] = 1

    data.loc[data[u'ОКБ (закалка)'] == 1.0, u'норм_1'] = 1
    data.loc[data[u'ОКБ (закалка)'] == 2.0, u'норм_2'] = 1
    data.loc[data[u'ОКБ (закалка)'] == 3.0, u'норм_3'] = 1
    data.loc[data[u'ОКБ (закалка)'] == 4.0, u'норм_4'] = 1

    data[u'длина (отпуск)'] = data[u'длина (отпуск)'].astype(float)
    data[u'длина (закалка)'] = data[u'длина (закалка)'].astype(float)
    data[u'скорость движения (норм.)'] = data[u'скорость движения (норм.)'].astype(float)
    data[u'скорость движения (отпуск)'] = data[u'скорость движения (отпуск)'].astype(float)
    data[u'расход воды (норм.)'] = data[u'расход воды (норм.) (1)'] + data[u'расход воды (норм.) (2)'] + data[
        u'расход воды (норм.) (3)']
    data[u'удельный расход воды'] = data[u'расход воды (норм.)'] * 1000.0 / (data[u'диаметр'] * np.pi)
    data[u'параметр отпуска'] = (data[u'темп-ра терм. (отпуск)'] + 273.0) \
                                * (
                                        20 + np.log(data[u'длина (отпуск)']) - np.log(
                                    data[u'скорость движения (отпуск)'] * 60.0)) \
                                * 1e-3
    data[u'параметр закалки'] = 1.0 / (1.0 / (data[u'темп-ра терм. (норм.)'] + 273.0) - 2.303 * 1.986 / 110000.0 * \
                                       (np.log10(data[u'длина (закалка)']) - \
                                        np.log10(data[u'скорость движения (норм.)'] * 60.0))) - 273.0
    data[u'Углеродный коэффицент'] = data[u'C'] + data[u'Mn'] / 6 + (data[u'Cr']) / 5 + (data[u'Ni'] + data[u'Cu']) / 15
    data[u'параметр отпуска'] = data[u'параметр отпуска'].apply(cl_inf)
    return data

# Функции, указывающие на нарушение границ признаков в input файле

def err_output(df, raw_name, left_b, right_b):
    all_df_len = df.shape[0]
    right_len = df[(df[raw_name]>left_b) & (df[raw_name]<right_b)].shape[0]
    err_df = pd.concat([df[df[raw_name]<=left_b],df[df[raw_name]>=right_b]])
    df_corr = df[(df[raw_name]>left_b) & (df[raw_name]<right_b)]
    err_df[u'Комментарий'] = u'нарушения границ %.1f<%s<%.1f'%(left_b, raw_name, right_b)
    err_df[u'Номер строки'] = err_df.index
    return [df_corr, err_df]

def less(df):
    start_shape = df.shape[0]
    df, err = err_output(df, u'диаметр', 0, 500)
    err_df = err
    df, err = err_output(df, u'толщина стенки', 0, 25)
    err_df = err_df.merge(err, how='outer')
    df, err = err_output(df, u'темп-ра терм. (норм.)', 500, 1000)
    err_df = err_df.merge(err, how='outer')
    df, err = err_output(df, u'темп-ра спрейр (норм.)', 500, 1000)
    err_df = err_df.merge(err, how='outer')
    df, err = err_output(df, u'скорость движения (норм.)', 0, 2.3)
    err_df = err_df.merge(err, how='outer')
    df, err = err_output(df, u'темп-ра терм. (отпуск)', 300, 800)
    err_df = err_df.merge(err, how='outer')
    df, err = err_output(df, u'темп-ра спрейр (отпуск)', 300, 800)
    err_df = err_df.merge(err, how='outer')
    df, err = err_output(df, u'скорость движения (отпуск)', 0, 2.3)
    err_df = err_df.merge(err, how='outer')
    return [df, err_df]

# Вызывает функции, представленные выше

def convert_df(df):
    df = auto_fill_empty(df)
    df, err_df = less(df)
    df = add_len_and_param(df)
    return [df, err_df]

ls_opt_need = [
    u'диаметр',
    u'толщина стенки',
    u'ОКБ (закалка)',
    u'ОКБ (отпуск)'
]
colomns = [u'Примечание',
    u'прочность',
    u'предел текучести',
    u'диаметр',
    u'толщина стенки',
    u'темп-ра терм. (норм.)',
    u'темп-ра спрейр (норм.)',
    u'скорость движения (норм.)',
    u'темп-ра терм. (отпуск)',
    u'темп-ра спрейр (отпуск)',
    u'скорость движения (отпуск)',
    u'ОКБ (закалка)',
    u'ОКБ (отпуск)',
    u'C',
    u'Mn',
    u'Si',
    u'Cr',
    u'Ni',
    u'Cu',
    u'Al',
    u'расход воды (норм.) (1)',
    u'расход воды (норм.) (2)',
    u'расход воды (норм.) (3)']


# Следущий набор функций ищет похожие режимы в исторических данных
# Находит сначала все данные с такой же толщиной стенки, затем с диаметром, затем с номером ОКБ закалки и отпуска
# Если нет точного совпадения по одному из признаков - берет ближайшее значение (функция: close_value)
# Затем беруться значения с максимальной скорость и максимальным номером партии (самые последнии режимы)
# Выводит все строчки с максимальной скоростью
# Если раскоментить строчку в функции max_value - будет выводить только одну строчку с максимальной скоростью (с самым большим номером парти)
# код не всегда самый красивый и оптимальный, но главное работает :)

def close_value(database, col, value):
    database[u'diff'] = np.abs(database[col] - value)
    return database[database[u'diff'] == min(database[u'diff'])][col].values[0]


def mod(data):
    data = data[colomns]
    return data


def max_value(row, val, i, database):
    tmp = database.copy()
    for col in ls_opt_need:
        zn = {
            'толщина стенки': row._3,
            'диаметр': row.диаметр,
            'ОКБ (закалка)': row._20,
            'ОКБ (отпуск)': row._21,
        }
        z = zn[col]
        tmp2 = tmp[tmp[col] == z]

        if tmp2.shape[0] == 0:
            tmp = tmp[tmp[col] == close_value(tmp, col, z)]
        else:
            tmp = tmp2.copy()
    if val == 'u':
        tmp[u'summ'] = tmp[u'скорость движения (норм.)'] + tmp[u'скорость движения (отпуск)']
        df = tmp[tmp[u'summ'] == tmp[u'summ'].max()]

        if df.shape[0] > 10:
            df = df[:10]

        df[u'Примечание'] = None
        df.loc[df[u'summ'] == df[u'summ'].max(), u'Примечание'] = 'Режим с максимальной скоростью ' + '(' + str(i) + ')'
        #         df = df[df['№ партии'] == df['№ партии'].max()]

        return df
    else:
        df = tmp[tmp[u'№ партии'] == tmp[u'№ партии'].max()]

        df = df[:1]

        df.loc[df[u'№ партии'] == tmp[
            u'№ партии'].max(), u'Примечание'] = 'Режим с максимальным номером партии ' + '(' + str(i) + ')'
        return df


def find(database, test):
    df1 = pd.DataFrame(columns=test.columns)
    df2 = pd.DataFrame(columns=test.columns)
    for row in test.itertuples():
        df1 = df1.append(max_value(row, 'u', row.Index + 1, database), sort=False)
        df2 = df2.append(max_value(row, 'p', row.Index + 1, database), sort=False)
        df1 = mod(df1)
        df2 = mod(df2)
    return df1, df2


# Чтобы выводилось все в цвете и было красиво
# Не самая красивая реализация, не было времени

def paint_over(first, end, color, sheet_ranges):
    st = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
          'W']
    for i in range(first, end + 1):
        for j in st:
            sheet_ranges[j + str(i)].fill = color


def color(direct, temp_conts, test, max_u):
    wb = load_workbook(direct)
    sheet_ranges = wb['Sheet1']
    color1 = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    color2 = PatternFill(start_color='00FF7F', end_color='00FF7F', fill_type='solid')
    color3 = PatternFill(start_color='00FF7F', end_color='00FF7F', fill_type='solid')
    k = 0
    for i in range(temp_conts):

        tmp = test[test[u'Примечание'] == 'Предсказание модели ' + '(' + str(i + 1) + ')']
        if tmp.shape[0] == 0:
            continue

        paint_over(2 + k, 2 + k, color1, sheet_ranges)
        tmp2 = max_u[max_u[u'Примечание'] == 'Режим с максимальной скоростью ' + '(' + str(i + 1) + ')']

        j = tmp2.shape[0]
        paint_over(3 + k, 2 + j + k, color2, sheet_ranges)
        paint_over(3 + j + k, 3 + j + k, color3, sheet_ranges)
        k = k + j + 2
    wb.save(direct)

def main(file, current_user):
    # Загруджаем перечень признаков, scaler и модели (градиентый бустинг и леса)

    titles_non_cat_data = json.load(open('app/model/titles_non_cat.json', "r"))
    titles_cat_data = json.load(open('app/model/titles_cat.json', "r"))

    scaler_fluidity = preprocessing.StandardScaler()
    scale_data_fluidity = json.load(open('app/model/fluidity/scaler', "r"))
    scaler_fluidity.mean_ = scale_data_fluidity[0]
    scaler_fluidity.scale_ = scale_data_fluidity[1]

    scaler_strength = preprocessing.StandardScaler()
    scale_data_strength = json.load(open('app/model/strength/scaler', "r"))
    scaler_strength.mean_ = scale_data_strength[0]
    scaler_strength.scale_ = scale_data_strength[1]

    grid_search_fluidity = pickle.load(open('app/model/fluidity/grid_search.sav', 'rb'))
    grid_search_strength = pickle.load(open('app/model/strength/grid_search.sav', 'rb'))

    GB_fluidity = pickle.load(open('app/model/fluidity/GB.sav', 'rb'))
    GB_strength = pickle.load(open('app/model/strength/GB.sav', 'rb'))

    # Загружаем входные и исторические данные
    input_data = pd.read_excel(file)
    now = datetime.now()
    time = "%d_%ddate %d_%d_%dtime" % (now.day, now.month, now.hour, now.minute, now.second)
    input_filename = os.getcwd() + '/app/INPUT/' + "prediction_input_" + time + "_" + current_user + ".xlsx"
    input_data.to_excel(input_filename)

    database = pd.read_excel('app/data/historical_data.xlsx')

    # Делаем отсев, если есть проблемы во входных данных
    test, err_test = convert_df(input_data)

    if test.shape[0] != 0:

        # Находим в исторических данных похожие режимы (с максимальной скоростью и номером партии)
        max_u, max_p = find(database, test)

        # Применяем scaler для нейронной сети (для остальных моделей в данном случае scaler не нужен!)
        sc_data = test[titles_non_cat_data]
        sc_data = scaler_fluidity.transform(sc_data)
        ct_data = test[titles_cat_data]
        sc_data = pd.DataFrame(sc_data, index=ct_data.index)
        inputs_fluidity = sc_data.combine_first(ct_data).values

        sc_data = test[titles_non_cat_data]
        sc_data = scaler_strength.transform(sc_data)
        ct_data = test[titles_cat_data]
        sc_data = pd.DataFrame(sc_data, index=ct_data.index)
        inputs_strength = sc_data.combine_first(ct_data).values

        # Загружаем нейронку
        json_file_fluidity = open('app/model/fluidity/NN.json', "r")
        loaded_model_fluidity = json_file_fluidity.read()
        json_file_fluidity.close()
        model_fluidity = model_from_json(loaded_model_fluidity)
        model_fluidity.load_weights('app/model/fluidity/NN.h5')
        model_fluidity.compile(loss=keras.losses.mean_squared_error, metrics=[keras.metrics.mean_squared_error],
                               optimizer=keras.optimizers.SGD(lr=0.0001, momentum=0.9, decay=1e-6))

        json_file_strength = open('app/model/strength/NN.json', "r")
        loaded_model_strength = json_file_strength.read()
        json_file_strength.close()
        model_strength = model_from_json(loaded_model_fluidity)
        model_strength.load_weights('app/model/strength/NN.h5')
        model_strength.compile(loss=keras.losses.mean_squared_error, metrics=[keras.metrics.mean_squared_error],
                               optimizer=keras.optimizers.SGD(lr=0.0001, momentum=0.9, decay=1e-6))

        # делаем предсказания для прочности и текучести
        # предсказываем значение по трем моделям и находим средний ответ (просто так точнее немного выходит)
        test[u'предел текучести'] = (model_fluidity.predict(inputs_fluidity)[:, 0] + grid_search_fluidity.predict(
            test[titles_non_cat_data + titles_cat_data]) + GB_fluidity.predict(
            test[titles_non_cat_data + titles_cat_data])) / 3
        test[u'прочность'] = (model_strength.predict(inputs_strength)[:, 0] + grid_search_strength.predict(
            test[titles_non_cat_data + titles_cat_data]) + GB_strength.predict(
            test[titles_non_cat_data + titles_cat_data])) / 3
        test[u'Примечание'] = None
        for i in range(input_data.shape[0]):
            test.loc[test.index == i, u'Примечание'] = 'Предсказание модели ' + '(' + str(i + 1) + ')'
        test = test[colomns]

    K.clear_session()

    # Сохраняем полученный результат
    out = pd.DataFrame(columns=colomns)
    error1 =""
    if test.shape[0] != 0:
        for i in range(input_data.shape[0]):
            out = out.append(test[test.index == i], sort=False)
            out = out.append(max_u[max_u[u'Примечание'] == 'Режим с максимальной скоростью ' + '(' + str(i + 1) + ')'],
                             sort=False)
            out = out.append(
                max_p[max_p[u'Примечание'] == 'Режим с максимальным номером партии ' + '(' + str(i + 1) + ')'],
                sort=False)

        direct = "app/OUTPUT/prediction_output" + time +  "_" + current_user +'.xlsx'
        writer = pd.ExcelWriter(direct, engine='xlsxwriter')
        out.to_excel(writer, index=False)
        sh = writer.sheets['Sheet1']
        sh.set_column(0, 0, 42)
        writer.close()
        wb = color(direct, input_data.shape[0], test, max_u)
    else:
        error1="Все строки удалены"


    cols = err_test.columns.tolist()
    cols = cols[-1:] + cols[:-1]
    err_test = err_test[cols]
    error2 = ""
    if err_test.shape[0] != 0:
        error2 = "Строка " + str(err_test['Номер строки'][0]) + " удалена из-за " + err_test['Комментарий'][0] + "; \n"
        for i in range(1, err_test.shape[0] - 1):
            error2 = error2 + "cтрока " + str(err_test['Номер строки'][i]) + " удалена из-за " + err_test['Комментарий'][i] + "; \n"
        if err_test.shape[0] > 1:
            error2 = error2 + "cтрока " + str(
                err_test['Номер строки'][err_test.shape[0] - 1]) + " удалена из-за " + err_test['Комментарий'][err_test.shape[0] - 1] + "\n"
    return error1, error2

